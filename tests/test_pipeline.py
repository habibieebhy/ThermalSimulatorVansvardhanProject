from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from mattress_intelligence.models import CompanyResearchRequest
from mattress_intelligence.pipeline import MattressIntelligencePipeline
from mattress_intelligence.settings import Settings


class PipelineTests(unittest.TestCase):
    def test_offline_demo_runs_end_to_end(self) -> None:
        project_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            settings = replace(
                Settings(),
                data_dir=root / "data",
                output_dir=root / "outputs",
                artifact_dir=root / "artifacts",
                database_path=root / "data" / "test.sqlite3",
                llm_provider="none",
            )
            pipeline = MattressIntelligencePipeline(settings)
            request = CompanyResearchRequest(
                company_name="Sleepwell Demo",
                official_domain="https://example.invalid/sleepwell",
                max_configurations_per_product=5,
                simulate_top_configurations=2,
            )
            output = root / "outputs" / "demo.xlsx"
            result = pipeline.import_catalogue(
                request, project_root / "examples" / "demo_catalogue.json", output
            )
            self.assertEqual(len(result.products), 3)
            self.assertTrue(result.configurations)
            self.assertTrue(output.exists())
            self.assertEqual(pipeline.repository.load(result.run_id).run_id, result.run_id)
            workbook = load_workbook(output, read_only=True)
            self.assertIn("Dashboard", workbook.sheetnames)
            self.assertIn("Configurations", workbook.sheetnames)
            self.assertIn("Discovery Log", workbook.sheetnames)
            self.assertIn("Evidence Sources", workbook.sheetnames)

    def test_imported_collection_can_skip_inference(self) -> None:
        project_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            settings = replace(
                Settings(),
                data_dir=root / "data",
                output_dir=root / "outputs",
                artifact_dir=root / "artifacts",
                database_path=root / "data" / "test.sqlite3",
                llm_provider="none",
                search_provider="none",
            )
            pipeline = MattressIntelligencePipeline(settings)
            request = CompanyResearchRequest(
                company_name="Collection Demo",
                official_domain="https://example.invalid/collection",
            )
            result = pipeline.import_catalogue(
                request,
                project_root / "examples" / "demo_catalogue.json",
                root / "outputs" / "collection.xlsx",
                analyze=False,
            )
            self.assertEqual(len(result.products), 3)
            self.assertFalse(result.configurations)
            self.assertFalse(result.simulations)
            self.assertTrue(any("Collection-only" in warning for warning in result.warnings))


if __name__ == "__main__":
    unittest.main()
